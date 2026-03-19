import os
from flask import Flask, send_from_directory, jsonify, request, field
import openpyxl
from datetime import (
   datetime,
)
app = Flask(__name__)


# caminho base do projeto (uma pasta acima do backend)
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))

# pasta frontend (html e js)
FRONTEND_DIR = os.path.join(BASE_DIR, "frontend")

#PASTA STATIC
STATIC_DIR = os.path.join(BASE_DIR, "static")

DB_DIR = os.path.join (os.path.dirname(__file__), "..", "db")
EXCEL_FILE = os.path.join (DB_DIR, "clienteS.xlSx")

COLUMNS = [
    "ID",
    "Nome",
    "CPF",
    "Email",
    "Telefone",
    "Endereço",
    "Observações",
    "Data Cadastro",

]

def init_excel():
    if not os.path.exists(DB_DIR):
       os.makedirs(DB_DIR)

    if not os.path.exists(EXCEL_FILE):
       Workbook = openpyxl.Workbook()
       sheet = Workbook.active
       sheet.title = "Clientes"
       sheet.append(COLUMNS)
       Workbook.save(EXCEL_FILE)

app = Flask(__name__, static_folder=STATIC_DIR,static_url_path="/"+ STATIC_DIR)


@app.route("/")
def home():
    return send_from_directory(FRONTEND_DIR,"index.html")


@app.route("/consulta")
def consultar_page():
    return send_from_directory(FRONTEND_DIR, "consulta.html")


@app.route("/alterar")
def alterar_page():
    return send_from_directory(FRONTEND_DIR, "alterar.html")

@app.route("assets/<path:filename>")
def assets(filename):
    return send_from_directory("../frontend/assets", filename)

@app.route("/cadastrar", methods=["POST"])
def cadastar_cliente():

    try:
        data = request.json

        required_fields = ["nome", "cpf", "email", "telefone", "endereço"]
        if not all(field in data[field] for field in required_fields):
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "todos os campos obrigatórios devem ser preenchidos"
                    }   
                ),
                400,
            )       
        workbook = openpyxl.load_workbook(EXCEL_FILE) 
        sheet = workbook. active
        last_id = 0
        if sheet.max_row > 1:
            last_id = sheet.cell(row=sheet.max_row, column=1). value or 0
        new_id = last_id + 1

        novo_cliente = [
            new_id,
            data.get("nome"),
            data.get("cpf"),
            data.get("email"),
            data.get("telefone"),
            data.get("endereço"),
            data.get("observaçoes", ""),
            datetime.now().strftime("%Y-%m-%d"),
        ]

        sheet.append(novo_cliente)
        workbook.save(EXCEL_FILE)

        return (
            jsonify(
                {
                    "status": "success",
                    "message": "Cliente cadastrado com sucesso!",
                    "id": new_id, 
                }
            ),
            201,
        )
    except Exception as e:
        
        return (
            jsonify({"status": "error", "message": f"Erro ao salvar no servidor: {e}"}),
            500,   
        )
        
              
            



if __name__ == "__main__":
    print("BASE_DIR", BASE_DIR)
    print("FRONTEND_DIR", FRONTEND_DIR)
    print("STATIC_DIR", STATIC_DIR)
    init_excel()
    app.run(debug=True)
