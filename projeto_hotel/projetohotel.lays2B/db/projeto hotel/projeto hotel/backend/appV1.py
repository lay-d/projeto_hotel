# =============================================================================
# ⚙️ backend/app.py
# =============================================================================
# 💡 Este é o coração do nosso sistema!
# Aqui criamos um pequeno SERVIDOR WEB com Flask, que:
# 1️⃣ Responde às requisições do navegador (frontend);
# 2️⃣ Manipula os dados enviados (cadastro, consulta, alteração);
# 3️⃣ Salva tudo dentro de um arquivo Excel (.xlsx), que funciona
#    como o "banco de dados" do sistema.
# =============================================================================

import os  # 📁 Biblioteca para lidar com diretórios e caminhos de arquivos
from flask import (
    Flask,
    request,
    jsonify,
    send_from_directory,
)  # 🌐 Importa o framework Flask e funções úteis
import openpyxl  # 📊 Biblioteca para ler e escrever planilhas Excel (.xlsx)
from datetime import (
    datetime,
)  # ⏰ Para registrar a data de cada cadastro automaticamente

# =============================================================================
# 🧠 INICIALIZAÇÃO DO SERVIDOR FLASK
# =============================================================================
# 💡 Aqui criamos a “instância” do Flask, que será o servidor web.
# - static_folder → indica onde estão os arquivos estáticos (CSS, JS, imagens)
# - static_url_path → define o caminho usado no navegador para acessá-los
# =============================================================================
app = Flask(__name__, static_folder="../static", static_url_path="/static")

# =============================================================================
# ⚙️ CONFIGURAÇÕES GERAIS E CONSTANTES
# =============================================================================
# 💾 Aqui definimos os diretórios e o nome do arquivo Excel que servirá de banco.
# =============================================================================
DB_DIR = os.path.join(
    os.path.dirname(__file__), "..", "db"
)  # Pasta onde ficará o banco de dados
EXCEL_FILE = os.path.join(
    DB_DIR, "clientes.xlsx"
)  # Caminho completo até o arquivo Excel

# Cabeçalhos das colunas do Excel (linha 1)
COLUMNS = [
    "ID",
    "Nome",
    "CPF",
    "Email",
    "Telefone",
    "Endereço",
    "Observações",
    "Data Cadastro",
]


# =============================================================================
# 🛠 FUNÇÃO AUXILIAR: CRIA O ARQUIVO EXCEL CASO NÃO EXISTA
# =============================================================================
# 💡 Essa função garante que o sistema funcione mesmo na primeira execução.
#    Se a pasta 'db' ou o arquivo 'clientes.xlsx' não existirem, eles são criados.
# =============================================================================
def init_excel():
    if not os.path.exists(DB_DIR):
        os.makedirs(DB_DIR)  # 🧱 Cria a pasta db se ainda não existir

    if not os.path.exists(EXCEL_FILE):
        workbook = openpyxl.Workbook()  # 📗 Cria uma nova planilha Excel
        sheet = workbook.active  # Pega a planilha ativa
        sheet.title = "Clientes"  # Nomeia a aba principal
        sheet.append(COLUMNS)  # Adiciona os títulos das colunas
        workbook.save(EXCEL_FILE)  # Salva o arquivo Excel


# =============================================================================
# 📂 ROTAS DE PÁGINAS HTML (FRONTEND)
# =============================================================================
# 💡 As rotas abaixo enviam os arquivos HTML para o navegador quando o
#    usuário acessa os endereços correspondentes.
# =============================================================================


@app.route("/")
def index():
    # 🏠 Página inicial (cadastro de cliente)
    return send_from_directory("../frontend", "index.html")


@app.route("/consulta")
def consulta_page():
    # 🔍 Página de consulta de clientes
    return send_from_directory("../frontend", "consulta.html")


@app.route("/alterar")
def alterar_page():
    # ✏️ Página de alteração de dados
    return send_from_directory("../frontend", "alterar.html")


# 📸 Rota para servir imagens, scripts ou outros arquivos na pasta “assets”
@app.route("/assets/<path:filename>")
def assets(filename):
    return send_from_directory("../frontend/assets", filename)


# =============================================================================
# 🚀 API (INTERFACE DE COMUNICAÇÃO BACKEND ↔ FRONTEND)
# =============================================================================
# 💡 “API” = conjunto de rotas que permitem que o site envie e receba dados
#    sem precisar recarregar a página inteira (usando JavaScript/JSON).
# =============================================================================


# -------------------------------------------------------------------------
# 📦 CADASTRAR CLIENTE
# -------------------------------------------------------------------------
@app.route("/api/cadastrar", methods=["POST"])
def cadastrar_cliente():
    """
    Recebe os dados do formulário (em JSON), valida e salva um novo cliente no Excel.
    """
    try:
        data = request.json  # 📨 Dados enviados do frontend via POST (JSON)

        # ⚙️ Campos obrigatórios que o usuário deve preencher
        required_fields = ["nome", "cpf", "email", "telefone", "endereco"]
        if not all(field in data and data[field] for field in required_fields):
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Todos os campos obrigatórios devem ser preenchidos.",
                    }
                ),
                400,
            )

        workbook = openpyxl.load_workbook(EXCEL_FILE)  # 📂 Abre o arquivo Excel
        sheet = workbook.active

        # 🧮 Cria um ID automático (último ID + 1)
        last_id = 0
        if sheet.max_row > 1:
            last_id = sheet.cell(row=sheet.max_row, column=1).value or 0
        new_id = last_id + 1

        # 🧾 Cria uma nova linha com os dados informados
        novo_cliente = [
            new_id,
            data.get("nome"),
            data.get("cpf"),
            data.get("email"),
            data.get("telefone"),
            data.get("endereco"),
            data.get("observacoes", ""),  # Campo opcional
            datetime.now().strftime("%Y-%m-%d"),  # 📅 Data atual
        ]

        sheet.append(novo_cliente)  # Adiciona nova linha no Excel
        workbook.save(EXCEL_FILE)  # 💾 Salva alterações

        # ✅ Retorna mensagem de sucesso
        return (
            jsonify(
                {
                    "status": "success",
                    "message": "Cliente cadastrado com sucesso!",
                    "id": new_id,
                }
            ),
            201,
        )

    except Exception as e:
        # ⚠️ Tratamento de erro genérico
        return (
            jsonify({"status": "error", "message": f"Erro ao salvar no servidor: {e}"}),
            500,
        )


# -------------------------------------------------------------------------
# 🔍 CONSULTAR CLIENTES
# -------------------------------------------------------------------------
@app.route("/api/buscar", methods=["GET"])
def buscar_clientes():
    """
    Busca clientes pelo nome (não diferencia maiúsculas/minúsculas).
    """
    nome_query = request.args.get("nome", "").lower()  # 🔤 Nome pesquisado

    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active
        resultados = []  # 🧺 Lista para armazenar resultados

        # 🧭 Percorre todas as linhas (ignorando o cabeçalho)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            cliente = dict(zip(COLUMNS, row))  # Converte linha → dicionário
            nome_cliente = (cliente.get("Nome") or "").lower()

            if nome_query in nome_cliente:
                resultados.append(cliente)

        return jsonify(resultados)  # 🔙 Retorna lista de clientes encontrados

    except FileNotFoundError:
        return (
            jsonify({"status": "error", "message": "Arquivo de dados não encontrado."}),
            404,
        )
    except Exception as e:
        return (
            jsonify({"status": "error", "message": f"Erro ao ler os dados: {e}"}),
            500,
        )


# -------------------------------------------------------------------------
# 📋 CONSULTAR CLIENTE POR ID
# -------------------------------------------------------------------------
@app.route("/api/cliente/<int:cliente_id>", methods=["GET"])
def get_cliente(cliente_id):
    """
    Retorna os dados completos de um cliente pelo seu ID.
    """
    try:
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active

        # 🔍 Procura o cliente linha por linha
        for row_idx in range(2, sheet.max_row + 1):
            row_id = sheet.cell(row=row_idx, column=1).value
            if row_id == cliente_id:
                row_values = [cell.value for cell in sheet[row_idx]]
                cliente = dict(zip(COLUMNS, row_values))
                return jsonify(cliente)

        # ❌ Se não encontrar o ID
        return jsonify({"status": "error", "message": "Cliente não encontrado."}), 404

    except Exception as e:
        return (
            jsonify({"status": "error", "message": f"Erro ao buscar cliente: {e}"}),
            500,
        )


# -------------------------------------------------------------------------
# ✏️ ATUALIZAR DADOS DE UM CLIENTE
# -------------------------------------------------------------------------
@app.route("/api/atualizar/<int:cliente_id>", methods=["POST"])
def atualizar_cliente(cliente_id):
    """
    Atualiza os dados de um cliente existente no Excel.
    """
    try:
        data = request.json
        workbook = openpyxl.load_workbook(EXCEL_FILE)
        sheet = workbook.active

        # 🧭 Encontra o cliente pelo ID
        row_to_update = -1
        for row_idx in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_idx, column=1).value == cliente_id:
                row_to_update = row_idx
                break

        if row_to_update == -1:
            return (
                jsonify(
                    {
                        "status": "error",
                        "message": "Cliente não encontrado para atualização.",
                    }
                ),
                404,
            )

        # 🧾 Atualiza as células com os novos dados
        sheet.cell(row=row_to_update, column=2, value=data.get("nome"))
        sheet.cell(row=row_to_update, column=3, value=data.get("cpf"))
        sheet.cell(row=row_to_update, column=4, value=data.get("email"))
        sheet.cell(row=row_to_update, column=5, value=data.get("telefone"))
        sheet.cell(row=row_to_update, column=6, value=data.get("endereco"))
        sheet.cell(row=row_to_update, column=7, value=data.get("observacoes"))

        workbook.save(EXCEL_FILE)  # 💾 Salva as alterações no Excel

        return jsonify(
            {
                "status": "success",
                "message": "Dados do cliente atualizados com sucesso!",
            }
        )

    except Exception as e:
        return (
            jsonify({"status": "error", "message": f"Erro ao atualizar dados: {e}"}),
            500,
        )


# =============================================================================
# 🏁 PONTO DE ENTRADA DA APLICAÇÃO
# =============================================================================
# 💡 Este trecho garante que o app só será executado se o arquivo for rodado
#    diretamente (python app.py), e não quando for importado.
# =============================================================================
if __name__ == "__main__":
    init_excel()  # ⚙️ Cria o Excel caso ainda não exista
    app.run(
        debug=True, port=5000
    )  # 🚀 Inicia o servidor local em http://localhost:5000
